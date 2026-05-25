import os
import csv
import io
import json
import re
import shutil
import tempfile
import threading
import time as _time
import traceback
import uuid
from datetime import date, datetime, timedelta

import requests
from dotenv import load_dotenv
from flask import Flask, Response, jsonify, render_template, request, send_file

load_dotenv()

app = Flask(__name__)

META_API_VERSION = os.getenv("META_API_VERSION", "v21.0")
META_ACCESS_TOKEN = os.getenv("META_ACCESS_TOKEN", "").strip()

GRAPH_URL = f"https://graph.facebook.com/{META_API_VERSION}"


PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
ACCOUNTS_FILE = os.path.join(PROJECT_DIR, "accounts.json")
TEMPLATES_DATA_DIR = os.path.join(PROJECT_DIR, "templates_data")

# 광고 계정별 템플릿 매핑. 키는 정규화된 act_xxx ID.
ACCOUNT_TEMPLATES = {
    "act_568276609290906": {
        "label": "British Council",
        "template_path": os.path.join(TEMPLATES_DATA_DIR, "bc_template.xlsx"),
        "raw_sheet": "raw_meta",
        "header_row": 1,
        # raw_meta 컬럼 매핑 (1-indexed):
        # A=date, B=campaign, C=adset, D=ad, E=impressions, F=link_clicks,
        # G=video_3sec, H=result_type, I=result_count, J=spend
        # K(creative)는 VLOOKUP 수식이 자동 채워짐 — 건드리지 않음
        "columns": {
            "date": 1,
            "campaign_name": 2,
            "adset_name": 3,
            "ad_name": 4,
            "impressions": 5,
            "link_clicks": 6,
            "video_3sec": 7,
            "result_type": 8,
            "result_count": 9,
            "spend": 10,
        },
    },
}


def _normalize_account_id(account_id: str) -> str:
    if not account_id:
        return ""
    account_id = account_id.strip()
    return account_id if account_id.startswith("act_") else f"act_{account_id}"


def _is_valid_account_format(account_id: str) -> bool:
    return account_id.startswith("act_") and account_id[4:].isdigit() and len(account_id) > 4


def _load_accounts_from_file():
    if not os.path.exists(ACCOUNTS_FILE):
        return None
    try:
        with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return [_normalize_account_id(x) for x in data.get("accounts", []) if x]
    except (json.JSONDecodeError, OSError):
        return []


def _save_accounts_to_file(ids):
    with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
        json.dump({"accounts": ids}, f, ensure_ascii=False, indent=2)


def _seed_from_env():
    raw = os.getenv("META_AD_ACCOUNT_IDS", "").strip() or os.getenv("META_AD_ACCOUNT_ID", "").strip()
    return [_normalize_account_id(x) for x in raw.split(",") if x.strip()]


def _initial_accounts():
    """accounts.json이 있으면 그걸 사용, 없으면 .env에서 시드하고 파일 생성."""
    from_file = _load_accounts_from_file()
    if from_file is not None:
        return from_file
    seeded = _seed_from_env()
    if seeded:
        _save_accounts_to_file(seeded)
    return seeded


ACCOUNT_IDS = _initial_accounts()
_account_meta_cache = {}

FIELDS = [
    "campaign_name",
    "adset_name",
    "ad_name",
    "impressions",
    "clicks",
    "spend",
    "cpc",
    "cpm",
    "ctr",
    "reach",
    "frequency",
    "actions",
]

CSV_COLUMNS = [
    "date_start",
    "date_stop",
    "campaign_name",
    "adset_name",
    "ad_name",
    "impressions",
    "clicks",
    "ctr",
    "video_views",
    "vtr",
    "conversions",
    "cvr",
    "spend",
]


def _resolve_account_id(requested):
    """요청 파라미터로 받은 account_id를 검증/정규화. 없으면 첫 번째 계정 사용."""
    if requested:
        normalized = _normalize_account_id(requested)
        if normalized in ACCOUNT_IDS:
            return normalized
        return None  # 허용되지 않은 ID
    return ACCOUNT_IDS[0] if ACCOUNT_IDS else None


def fetch_account_meta(account_id: str):
    """광고 계정의 이름/통화/타임존 등 메타정보 조회. 메모리 캐시 사용."""
    if account_id in _account_meta_cache:
        return _account_meta_cache[account_id]

    fallback = {"id": account_id, "name": account_id}
    if not META_ACCESS_TOKEN:
        return fallback

    try:
        r = requests.get(
            f"{GRAPH_URL}/{account_id}",
            params={
                "fields": "name,currency,timezone_name,account_status",
                "access_token": META_ACCESS_TOKEN,
            },
            timeout=15,
        )
        data = r.json()
    except requests.RequestException:
        return fallback

    if "error" in data:
        meta = {**fallback, "error": data["error"].get("message")}
    else:
        meta = {
            "id": account_id,
            "name": data.get("name", account_id),
            "currency": data.get("currency"),
            "timezone": data.get("timezone_name"),
            "status": data.get("account_status"),
        }
    _account_meta_cache[account_id] = meta
    return meta


def _parse_target(campaign_name):
    """campaign_name → 'Adults' / 'Young Learners' / 'Kinder' / '기타' (프론트 parseTarget 동일)"""
    c = (campaign_name or "").lower()
    if "kinder" in c:
        return "Kinder"
    if "adults" in c:
        return "Adults"
    if "younglearners" in c or "youngleaners" in c:
        return "Young Learners"
    if "parents" in c:
        return "Kinder"
    return "기타"


def _parse_targeting_main(campaign_name, adset_name):
    """메인 Targeting tag만 (수정자 제외) — Creative 추출용. 'A+'/'LAL'/'PT'/'Ilsan'/None"""
    src = ((adset_name or "") + " " + (campaign_name or "")).lower()
    if re.search(r"\ba\+|aplus", src):
        return "A+"
    if re.search(r"\blal", src):
        return "LAL"
    if re.search(r"prospect|\bpt\b", src):
        return "PT"
    if "ilsan" in src:
        return "Ilsan"
    return None


def _parse_targeting(campaign_name, adset_name):
    """메인 tag + 수정자 결합 (프론트 parseTargeting 동일)"""
    main = _parse_targeting_main(campaign_name, adset_name)
    if not main:
        last = (adset_name or "").split("-")
        return last[-1] if last and last[-1] else "기타"
    src = ((adset_name or "") + " " + (campaign_name or "")).lower()
    mods = []
    if "primary" in src:
        mods.append("Primary")
    if "summer" in src:
        mods.append("summer")
    if "winter" in src:
        mods.append("winter")
    if re.search(r"\bbau\b", src) and not mods:
        mods.append("BAU")
    return f"{main} - {', '.join(mods)}" if mods else main


def _parse_creative(ad_name, main_tag):
    """ad_name에서 메인 tag 뒤 부분 짧은 라벨로 추출. 못 찾으면 마지막 segment."""
    if not ad_name:
        return ""
    if not main_tag:
        parts = ad_name.split("-")
        return parts[-1] if parts else ad_name
    pattern = re.escape(main_tag)
    m = re.search(pattern, ad_name, re.IGNORECASE)
    if not m:
        parts = ad_name.split("-")
        return parts[-1] if parts else ad_name
    after = ad_name[m.end():]
    after = re.sub(r"^[-_]+", "", after)
    return after or ad_name


def _annotate_row_for_export(row):
    """행에 _target / _targeting / _targeting_main / _creative 부여."""
    row["_target"] = _parse_target(row.get("campaign_name"))
    main = _parse_targeting_main(row.get("campaign_name"), row.get("adset_name"))
    row["_targeting_main"] = main or "기타"
    row["_targeting"] = _parse_targeting(row.get("campaign_name"), row.get("adset_name"))
    row["_creative"] = _parse_creative(row.get("ad_name"), main) or (row.get("ad_name") or "")
    return row


def _enrich_row(row):
    """Add video_views/vtr/conversions/cvr fields based on actions array.

    - video_views: None if no video_view action (이미지 소재 신호)
    - vtr: None if video_views is None
    - conversions: 0 if no conversion-type action found
    - cvr: 0 if no impressions or no conversions
    """
    actions = row.get("actions") or []
    try:
        impressions = float(row.get("impressions") or 0)
    except (TypeError, ValueError):
        impressions = 0.0

    video_views = None
    for a in actions:
        if a.get("action_type") == "video_view":
            try:
                video_views = float(a.get("value") or 0)
            except (TypeError, ValueError):
                video_views = 0.0
            break

    if video_views is not None:
        row["video_views"] = int(video_views)
        row["vtr"] = (video_views / impressions * 100) if impressions else 0.0
    else:
        row["video_views"] = None
        row["vtr"] = None

    conv = 0.0
    for at in _CONVERSION_ACTION_TYPES:
        v = _extract_action_value(actions, at)
        if v:
            conv = v
            break
    row["conversions"] = int(conv)
    row["cvr"] = (conv / impressions * 100) if (impressions and conv) else 0.0
    return row


def fetch_insights(date_preset=None, since=None, until=None, level="campaign", account_id=None):
    if not META_ACCESS_TOKEN or not ACCOUNT_IDS:
        return {
            "error": "META_ACCESS_TOKEN 또는 META_AD_ACCOUNT_IDS가 비어 있습니다. .env 파일을 채워주세요.",
            "code": "missing_credentials",
        }

    resolved = _resolve_account_id(account_id)
    if not resolved:
        return {
            "error": "허용되지 않은 광고 계정 ID입니다. .env의 META_AD_ACCOUNT_IDS에 추가해주세요.",
            "code": "invalid_account",
        }
    url = f"{GRAPH_URL}/{resolved}/insights"

    params = {
        "fields": ",".join(FIELDS),
        "level": level,
        "access_token": META_ACCESS_TOKEN,
        "limit": 200,
    }

    # 광고 단위는 항상 일자별 분해 (행: 날짜 × 광고)
    if level == "ad":
        params["time_increment"] = 1

    if since and until:
        params["time_range"] = json.dumps({"since": since, "until": until})
    else:
        params["date_preset"] = date_preset or "last_7d"

    # 페이지네이션 처리
    all_rows = []
    next_url, next_params = url, params
    pages = 0
    while next_url and pages < 50:
        pages += 1
        try:
            r = requests.get(next_url, params=next_params, timeout=30)
            data = r.json()
        except requests.RequestException as e:
            return {"error": f"네트워크 오류: {e}"}
        if "error" in data:
            err = data["error"]
            return {
                "error": err.get("message", "Meta API 오류"),
                "code": err.get("code"),
                "type": err.get("type"),
            }
        all_rows.extend(data.get("data", []))
        paging = data.get("paging", {})
        next_url = paging.get("next")
        next_params = None

    enriched = [_enrich_row(r) for r in all_rows]
    return {"data": enriched}


def _get_week_ranges(today=None):
    """이번 주(월요일~오늘) vs 같은 길이의 지난 주 (지난 주 월요일~동일 요일).

    오늘이 월요일이면 1일치 vs 1일치 비교가 되어 공정한 비교가 됩니다.
    """
    today = today or date.today()
    days_since_monday = today.weekday()  # Mon=0, Sun=6
    this_start = today - timedelta(days=days_since_monday)
    this_end = today
    last_start = this_start - timedelta(days=7)
    last_end = last_start + timedelta(days=days_since_monday)
    return this_start, this_end, last_start, last_end


def _to_float(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _aggregate(rows):
    impressions = sum(_to_float(r.get("impressions")) for r in rows)
    clicks = sum(_to_float(r.get("clicks")) for r in rows)
    spend = sum(_to_float(r.get("spend")) for r in rows)
    return {
        "impressions": impressions,
        "clicks": clicks,
        "spend": spend,
        "ctr": (clicks / impressions * 100) if impressions else 0.0,
        "cpc": (spend / clicks) if clicks else 0.0,
        "cpm": (spend / impressions * 1000) if impressions else 0.0,
    }


def _pct_change(now, prev):
    if not prev:
        return None
    return (now - prev) / prev * 100


def _build_deltas(this_total, last_total):
    out = {}
    for k in ("spend", "impressions", "clicks", "ctr", "cpc", "cpm"):
        out[k] = {
            "this": this_total[k],
            "last": last_total[k],
            "abs": this_total[k] - last_total[k],
            "pct": _pct_change(this_total[k], last_total[k]),
        }
    return out


def _build_movers(this_rows, last_rows, top_n=3):
    this_by = {r.get("campaign_name", ""): r for r in this_rows}
    last_by = {r.get("campaign_name", ""): r for r in last_rows}
    names = set(this_by) | set(last_by)
    movers = []
    for name in names:
        if not name:
            continue
        this_spend = _to_float(this_by.get(name, {}).get("spend"))
        last_spend = _to_float(last_by.get(name, {}).get("spend"))
        movers.append({
            "campaign_name": name,
            "this_spend": this_spend,
            "last_spend": last_spend,
            "delta": this_spend - last_spend,
            "delta_pct": _pct_change(this_spend, last_spend),
        })
    movers.sort(key=lambda x: x["delta"], reverse=True)
    up = [m for m in movers if m["delta"] > 0][:top_n]
    down = [m for m in reversed(movers) if m["delta"] < 0][:top_n]
    return up, down


def _fmt_won(v):
    return f"{int(round(_to_float(v))):,}원"


def _fmt_int(v):
    return f"{int(round(_to_float(v))):,}"


def _fmt_pct_delta(pct):
    if pct is None:
        return "비교 불가"
    arrow = "▲" if pct > 0 else ("▼" if pct < 0 else "—")
    return f"{arrow} {abs(pct):.1f}%"


def _build_bullets(this_total, last_total, deltas, up, down, period):
    bullets = []

    sp = deltas["spend"]
    direction = "증가" if sp["abs"] > 0 else ("감소" if sp["abs"] < 0 else "유지")
    bullets.append(
        f"이번 주({period['this']['start']}~{period['this']['end']}) 총 지출은 "
        f"{_fmt_won(sp['this'])}로 지난 주({_fmt_won(sp['last'])}) 대비 "
        f"{_fmt_pct_delta(sp['pct'])} {direction}했습니다."
    )

    cl = deltas["clicks"]
    imp = deltas["impressions"]
    bullets.append(
        f"노출 {_fmt_int(imp['this'])}회 ({_fmt_pct_delta(imp['pct'])}), "
        f"클릭 {_fmt_int(cl['this'])}회 ({_fmt_pct_delta(cl['pct'])})"
    )

    ctr = deltas["ctr"]
    cpc = deltas["cpc"]
    bullets.append(
        f"효율: CTR {ctr['this']:.2f}% ({_fmt_pct_delta(ctr['pct'])}), "
        f"CPC {_fmt_won(cpc['this'])} ({_fmt_pct_delta(cpc['pct'])})"
    )

    if up:
        m = up[0]
        bullets.append(
            f"지출 증가 1위 캠페인: \"{m['campaign_name']}\" — "
            f"{_fmt_won(m['last_spend'])} → {_fmt_won(m['this_spend'])} "
            f"({_fmt_pct_delta(m['delta_pct'])})"
        )

    if down:
        m = down[0]
        bullets.append(
            f"지출 감소 1위 캠페인: \"{m['campaign_name']}\" — "
            f"{_fmt_won(m['last_spend'])} → {_fmt_won(m['this_spend'])} "
            f"({_fmt_pct_delta(m['delta_pct'])})"
        )

    return bullets


@app.route("/")
def index():
    token_ready = bool(META_ACCESS_TOKEN)
    return render_template(
        "index.html",
        credentials_ready=token_ready,
        has_accounts=bool(ACCOUNT_IDS),
    )


@app.route("/api/accounts", methods=["GET"])
def api_accounts_list():
    accounts = []
    for aid in ACCOUNT_IDS:
        meta = dict(fetch_account_meta(aid))
        tpl = ACCOUNT_TEMPLATES.get(aid)
        meta["template"] = {"label": tpl["label"]} if tpl else None
        accounts.append(meta)
    return jsonify({
        "accounts": accounts,
        "default": ACCOUNT_IDS[0] if ACCOUNT_IDS else None,
    })


@app.route("/api/accounts", methods=["POST"])
def api_accounts_add():
    if not META_ACCESS_TOKEN:
        return jsonify({"error": "META_ACCESS_TOKEN이 비어 있습니다. .env를 확인하세요."}), 400

    body = request.get_json(silent=True) or {}
    raw = (body.get("id") or "").strip()
    if not raw:
        return jsonify({"error": "광고 계정 ID를 입력해주세요."}), 400

    normalized = _normalize_account_id(raw)
    if not _is_valid_account_format(normalized):
        return jsonify({"error": f"올바른 형식이 아닙니다: '{normalized}'. 숫자만 또는 act_숫자 형태여야 합니다."}), 400

    if normalized in ACCOUNT_IDS:
        return jsonify({"error": "이미 등록된 계정입니다.", "id": normalized}), 409

    # Meta API 호출로 토큰 권한 + 계정 존재 검증
    try:
        r = requests.get(
            f"{GRAPH_URL}/{normalized}",
            params={
                "fields": "name,currency,timezone_name,account_status",
                "access_token": META_ACCESS_TOKEN,
            },
            timeout=15,
        )
        data = r.json()
    except requests.RequestException as e:
        return jsonify({"error": f"네트워크 오류: {e}"}), 400

    if "error" in data:
        msg = data["error"].get("message", "Meta API 오류")
        return jsonify({"error": f"검증 실패: {msg}"}), 400

    ACCOUNT_IDS.append(normalized)
    _save_accounts_to_file(ACCOUNT_IDS)
    _account_meta_cache.pop(normalized, None)
    meta = fetch_account_meta(normalized)
    return jsonify({"ok": True, "account": meta}), 201


@app.route("/api/accounts/<account_id>", methods=["DELETE"])
def api_accounts_delete(account_id):
    normalized = _normalize_account_id(account_id)
    if normalized not in ACCOUNT_IDS:
        return jsonify({"error": "등록되지 않은 계정입니다."}), 404
    ACCOUNT_IDS.remove(normalized)
    _save_accounts_to_file(ACCOUNT_IDS)
    _account_meta_cache.pop(normalized, None)
    return jsonify({"ok": True, "id": normalized})


@app.route("/api/weekly-insights")
def api_weekly_insights():
    account_id = request.args.get("account_id")
    this_start, this_end, last_start, last_end = _get_week_ranges()

    this_result = fetch_insights(
        since=this_start.isoformat(),
        until=this_end.isoformat(),
        level="campaign",
        account_id=account_id,
    )
    if "error" in this_result:
        return jsonify(this_result), 400

    last_result = fetch_insights(
        since=last_start.isoformat(),
        until=last_end.isoformat(),
        level="campaign",
        account_id=account_id,
    )
    if "error" in last_result:
        return jsonify(last_result), 400

    this_rows = this_result["data"]
    last_rows = last_result["data"]

    this_total = _aggregate(this_rows)
    last_total = _aggregate(last_rows)
    deltas = _build_deltas(this_total, last_total)
    up, down = _build_movers(this_rows, last_rows)

    period = {
        "this": {"start": this_start.isoformat(), "end": this_end.isoformat()},
        "last": {"start": last_start.isoformat(), "end": last_end.isoformat()},
        "days": (this_end - this_start).days + 1,
    }

    bullets = _build_bullets(this_total, last_total, deltas, up, down, period)

    return jsonify({
        "period": period,
        "totals": {"this": this_total, "last": last_total},
        "deltas": deltas,
        "top_spend_up": up,
        "top_spend_down": down,
        "bullets": bullets,
    })


@app.route("/api/comparison")
def api_comparison():
    """두 사용자 정의 기간 간 비교 데이터.

    Query params:
      account_id (필수)
      current_start, current_end (필수, YYYY-MM-DD)
      previous_start, previous_end (필수, YYYY-MM-DD)

    응답:
      {
        current: { period, rows: [광고 단위 일별 행, _enrich_row 적용됨] },
        previous: { period, rows: [...] }
      }
    필터링/집계는 클라이언트가 처리.
    """
    if not META_ACCESS_TOKEN or not ACCOUNT_IDS:
        return jsonify({"error": "자격 증명 미설정"}), 400
    resolved = _resolve_account_id(request.args.get("account_id"))
    if not resolved:
        return jsonify({"error": "허용되지 않은 광고 계정 ID입니다."}), 400

    cs = request.args.get("current_start")
    ce = request.args.get("current_end")
    ps = request.args.get("previous_start")
    pe = request.args.get("previous_end")
    if not all([cs, ce, ps, pe]):
        return jsonify({
            "error": "current_start, current_end, previous_start, previous_end 모두 필요합니다."
        }), 400

    cur_res = fetch_insights(since=cs, until=ce, level="ad", account_id=resolved)
    if "error" in cur_res:
        return jsonify(cur_res), 400
    prev_res = fetch_insights(since=ps, until=pe, level="ad", account_id=resolved)
    if "error" in prev_res:
        return jsonify(prev_res), 400

    def _days(s, e):
        try:
            d1 = datetime.strptime(s, "%Y-%m-%d").date()
            d2 = datetime.strptime(e, "%Y-%m-%d").date()
            return (d2 - d1).days + 1
        except (ValueError, TypeError):
            return None

    return jsonify({
        "current": {
            "period": {"start": cs, "end": ce, "days": _days(cs, ce)},
            "rows": cur_res["data"],
        },
        "previous": {
            "period": {"start": ps, "end": pe, "days": _days(ps, pe)},
            "rows": prev_res["data"],
        },
    })


@app.route("/api/insights")
def api_insights():
    result = fetch_insights(
        date_preset=request.args.get("date_preset"),
        since=request.args.get("since"),
        until=request.args.get("until"),
        level=request.args.get("level", "campaign"),
        account_id=request.args.get("account_id"),
    )
    status = 400 if "error" in result else 200
    return jsonify(result), status


def _paged_get(url, params):
    """페이지네이션을 순회하며 모든 data 행을 수집."""
    all_rows = []
    next_url, next_params = url, params
    pages = 0
    while next_url and pages < 50:
        pages += 1
        try:
            r = requests.get(next_url, params=next_params, timeout=60)
            data = r.json()
        except requests.RequestException as e:
            return {"error": f"네트워크 오류: {e}"}
        if "error" in data:
            return {"error": data["error"].get("message", "Meta API 오류")}
        all_rows.extend(data.get("data", []))
        paging = data.get("paging", {})
        next_url = paging.get("next")
        next_params = None
    return {"data": all_rows}


def _fetch_campaigns_meta(account_id):
    return _paged_get(
        f"{GRAPH_URL}/{account_id}/campaigns",
        {
            "fields": "id,name,status,effective_status,start_time,stop_time,daily_budget,lifetime_budget,objective",
            "limit": 200,
            "access_token": META_ACCESS_TOKEN,
        },
    )


def _fetch_cumulative_spend(account_id, earliest_start, until):
    """모든 캠페인의 누적 spend를 한 번에 조회."""
    res = _paged_get(
        f"{GRAPH_URL}/{account_id}/insights",
        {
            "fields": "campaign_id,spend",
            "level": "campaign",
            "time_range": json.dumps({"since": earliest_start, "until": until}),
            "limit": 500,
            "access_token": META_ACCESS_TOKEN,
        },
    )
    if "error" in res:
        return res
    by_camp = {}
    for row in res["data"]:
        cid = row.get("campaign_id")
        if cid:
            by_camp[cid] = by_camp.get(cid, 0.0) + _to_float(row.get("spend"))
    return {"data": by_camp}


def _fetch_period_ads_for_breakdown(account_id, since, until):
    """기간 내 광고 단위 데이터 (계층 빌드용). time_increment 없음 — 기간 합산."""
    res = _paged_get(
        f"{GRAPH_URL}/{account_id}/insights",
        {
            "fields": "campaign_id,campaign_name,adset_id,adset_name,ad_id,ad_name,impressions,clicks,spend,ctr,actions",
            "level": "ad",
            "time_range": json.dumps({"since": since, "until": until}),
            "limit": 500,
            "access_token": META_ACCESS_TOKEN,
        },
    )
    if "error" in res:
        return res
    return {"data": [_enrich_row(r) for r in res["data"]]}


def _parse_iso_date(s):
    if not s:
        return None
    try:
        # ISO 8601 with timezone (e.g. "2026-04-01T00:00:00+0900") or just date
        s_norm = s.split("T")[0]
        return datetime.strptime(s_norm, "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        return None


def _calc_progress(start_str, stop_str, daily_budget, lifetime_budget, cumul_spend, until_str):
    out = {
        "time_pct": None,
        "spend_pct": None,
        "cumul_spend": cumul_spend,
        "total_budget": None,
    }
    start_d = _parse_iso_date(start_str)
    stop_d = _parse_iso_date(stop_str)
    until_d = _parse_iso_date(until_str) or date.today()

    if start_d and stop_d and stop_d > start_d:
        total_days = (stop_d - start_d).days
        effective_end = min(until_d, stop_d)
        elapsed = max(0, (effective_end - start_d).days)
        out["time_pct"] = max(0.0, min(100.0, elapsed / total_days * 100))

    lb = _to_float(lifetime_budget) if lifetime_budget else 0
    db = _to_float(daily_budget) if daily_budget else 0

    if lb:
        out["total_budget"] = lb
    elif db and start_d and stop_d and stop_d > start_d:
        out["total_budget"] = db * (stop_d - start_d).days

    if out["total_budget"]:
        out["spend_pct"] = cumul_spend / out["total_budget"] * 100

    return out


def _aggregate_enriched(rows):
    """_enrich_row 적용된 행들을 집계."""
    imps = sum(_to_float(r.get("impressions")) for r in rows)
    clicks = sum(_to_float(r.get("clicks")) for r in rows)
    spend = sum(_to_float(r.get("spend")) for r in rows)

    has_video = any(r.get("video_views") is not None for r in rows)
    if has_video:
        video_views = sum(_to_float(r.get("video_views") or 0) for r in rows if r.get("video_views") is not None)
        vtr = (video_views / imps * 100) if imps else 0.0
    else:
        video_views = None
        vtr = None

    conv = sum(_to_float(r.get("conversions") or 0) for r in rows)

    return {
        "impressions": imps,
        "clicks": clicks,
        "ctr": (clicks / imps * 100) if imps else 0.0,
        "video_views": int(video_views) if (video_views is not None) else None,
        "vtr": vtr,
        "conversions": int(conv),
        "cvr": (conv / imps * 100) if (imps and conv) else 0.0,
        "spend": spend,
    }


def _parse_overrides(s):
    """override 쿼리 파라미터 파싱: 'cid1:start1:stop1,cid2:start2:stop2' → dict."""
    if not s:
        return {}
    out = {}
    for part in s.split(","):
        bits = part.strip().split(":")
        if len(bits) == 3:
            cid, start, stop = bits
            if cid and start and stop:
                out[cid.strip()] = {"start": start.strip(), "stop": stop.strip()}
    return out


def _build_breakdown(resolved, since, until, overrides):
    """캠페인×광고세트×광고 계층 + 진행률 계산."""
    camp_res = _fetch_campaigns_meta(resolved)
    if "error" in camp_res:
        return camp_res
    campaigns = camp_res["data"]

    if not campaigns:
        return {"period": {"since": since, "until": until}, "campaigns": []}

    # 누적 spend 조회용 earliest start 결정
    starts = []
    for c in campaigns:
        d = _parse_iso_date(c.get("start_time"))
        if d:
            starts.append(d)
    earliest_d = min(starts) if starts else _parse_iso_date(until)
    earliest_str = earliest_d.isoformat() if earliest_d else until

    cumul_res = _fetch_cumulative_spend(resolved, earliest_str, until)
    if "error" in cumul_res:
        return cumul_res
    cumul_by_camp = cumul_res["data"]

    ad_res = _fetch_period_ads_for_breakdown(resolved, since, until)
    if "error" in ad_res:
        return ad_res
    ad_rows = ad_res["data"]

    # campaign_id → adset_id → ads list
    by_camp = {}
    for r in ad_rows:
        cid = r.get("campaign_id")
        if not cid:
            continue
        c_entry = by_camp.setdefault(cid, {"adsets": {}})
        aset_id = r.get("adset_id")
        a_entry = c_entry["adsets"].setdefault(aset_id, {
            "id": aset_id,
            "name": r.get("adset_name", ""),
            "ads": [],
        })
        a_entry["ads"].append({
            "id": r.get("ad_id"),
            "name": r.get("ad_name", ""),
            "metrics": {
                "impressions": _to_float(r.get("impressions")),
                "clicks": _to_float(r.get("clicks")),
                "ctr": _to_float(r.get("ctr")),
                "video_views": r.get("video_views"),
                "vtr": r.get("vtr"),
                "conversions": r.get("conversions", 0),
                "cvr": r.get("cvr", 0),
                "spend": _to_float(r.get("spend")),
            },
        })

    out_campaigns = []
    for c in campaigns:
        cid = c["id"]
        group = by_camp.get(cid)
        if not group:
            continue

        for aset in group["adsets"].values():
            ad_rows_for_aggr = [{
                "impressions": ad["metrics"]["impressions"],
                "clicks": ad["metrics"]["clicks"],
                "spend": ad["metrics"]["spend"],
                "video_views": ad["metrics"]["video_views"],
                "vtr": ad["metrics"]["vtr"],
                "conversions": ad["metrics"]["conversions"],
            } for ad in aset["ads"]]
            aset["metrics"] = _aggregate_enriched(ad_rows_for_aggr)

        all_ad_metrics = []
        for aset in group["adsets"].values():
            for ad in aset["ads"]:
                all_ad_metrics.append({
                    "impressions": ad["metrics"]["impressions"],
                    "clicks": ad["metrics"]["clicks"],
                    "spend": ad["metrics"]["spend"],
                    "video_views": ad["metrics"]["video_views"],
                    "vtr": ad["metrics"]["vtr"],
                    "conversions": ad["metrics"]["conversions"],
                })
        campaign_metrics = _aggregate_enriched(all_ad_metrics)

        ov = overrides.get(cid) if overrides else None
        if ov:
            start_for_progress = ov["start"]
            stop_for_progress = ov["stop"]
        else:
            start_for_progress = c.get("start_time")
            stop_for_progress = c.get("stop_time")

        progress = _calc_progress(
            start_for_progress,
            stop_for_progress,
            c.get("daily_budget"),
            c.get("lifetime_budget"),
            cumul_by_camp.get(cid, 0.0),
            until,
        )

        out_campaigns.append({
            "id": cid,
            "name": c.get("name"),
            "status": c.get("effective_status") or c.get("status"),
            "objective": c.get("objective"),
            "schedule": {
                "start": c.get("start_time"),
                "stop": c.get("stop_time"),
                "override": ov,
            },
            "budget": {
                "daily": _to_float(c.get("daily_budget")) if c.get("daily_budget") else None,
                "lifetime": _to_float(c.get("lifetime_budget")) if c.get("lifetime_budget") else None,
                "total": progress.get("total_budget"),
            },
            "progress": {
                "time_pct": progress.get("time_pct"),
                "spend_pct": progress.get("spend_pct"),
                "cumul_spend": progress.get("cumul_spend"),
            },
            "metrics": campaign_metrics,
            "adsets": list(group["adsets"].values()),
        })

    out_campaigns.sort(key=lambda x: x["metrics"]["spend"], reverse=True)
    return {"period": {"since": since, "until": until}, "campaigns": out_campaigns}


@app.route("/api/breakdown")
def api_breakdown():
    if not META_ACCESS_TOKEN or not ACCOUNT_IDS:
        return jsonify({"error": "자격 증명 미설정"}), 400
    resolved = _resolve_account_id(request.args.get("account_id"))
    if not resolved:
        return jsonify({"error": "허용되지 않은 광고 계정 ID입니다."}), 400
    since, until = _resolve_date_range(
        request.args.get("since"),
        request.args.get("until"),
        request.args.get("date_preset"),
    )
    overrides = _parse_overrides(request.args.get("override", ""))
    result = _build_breakdown(resolved, since, until, overrides)
    status = 400 if "error" in result else 200
    return jsonify(result), status


def _fetch_template_insights(account_id, since, until):
    """템플릿 출력용 — 광고 단위, 일별 분해, actions(video_view 포함) 메트릭 포함."""
    template_fields = [
        "date_start",
        "campaign_name",
        "adset_name",
        "ad_name",
        "impressions",
        "inline_link_clicks",
        "spend",
        "actions",
    ]

    if not META_ACCESS_TOKEN:
        return {"error": "META_ACCESS_TOKEN이 비어 있습니다."}

    url = f"{GRAPH_URL}/{account_id}/insights"
    params = {
        "fields": ",".join(template_fields),
        "level": "ad",
        "time_increment": 1,
        "time_range": json.dumps({"since": since, "until": until}),
        "access_token": META_ACCESS_TOKEN,
        "limit": 500,
    }

    all_rows = []
    next_url = url
    next_params = params
    pages = 0
    while next_url and pages < 50:
        try:
            r = requests.get(next_url, params=next_params, timeout=60)
            data = r.json()
        except requests.RequestException as e:
            return {"error": f"네트워크 오류: {e}"}
        if "error" in data:
            return {"error": data["error"].get("message", "Meta API 오류")}
        all_rows.extend(data.get("data", []))
        paging = data.get("paging", {})
        next_url = paging.get("next")
        next_params = None  # next URL에 이미 모든 파라미터 포함됨
        pages += 1
    return {"data": all_rows}


# 결과(전환)로 카운트할 action_type 우선순위
_CONVERSION_ACTION_TYPES = [
    "offsite_conversion.fb_pixel_lead",
    "lead",
    "leadgen.other",
    "offsite_conversion.fb_pixel_complete_registration",
    "offsite_conversion.fb_pixel_purchase",
    "purchase",
]

# 광고 목표 → 한국어 결과 유형 표시
_RESULT_TYPE_LABELS = {
    "lead": "웹사이트 잠재 고객",
    "offsite_conversion.fb_pixel_lead": "웹사이트 잠재 고객",
    "leadgen.other": "잠재 고객",
    "offsite_conversion.fb_pixel_complete_registration": "등록 완료",
    "offsite_conversion.fb_pixel_purchase": "구매",
    "purchase": "구매",
}


def _extract_action_value(actions, action_type):
    if not actions:
        return 0
    for a in actions:
        if a.get("action_type") == action_type:
            try:
                return float(a.get("value") or 0)
            except (TypeError, ValueError):
                return 0
    return 0


def _summarize_result(actions):
    """Meta API actions 배열에서 (result_type_label, result_count) 추출."""
    if not actions:
        return ("", 0)
    for at in _CONVERSION_ACTION_TYPES:
        v = _extract_action_value(actions, at)
        if v:
            return (_RESULT_TYPE_LABELS.get(at, at), int(v))
    return ("", 0)


def _summarize_video_3sec(actions):
    """actions 배열에서 video_view (= 3초 이상 동영상 재생) 추출."""
    return int(_extract_action_value(actions, "video_view"))


# ===== 백그라운드 잡 매니저 (Excel 리포트 생성용) =====
_jobs = {}
_jobs_lock = threading.Lock()
JOB_TTL_SECONDS = 1800  # 30분


def _emit(job_id, msg):
    """터미널 + 잡 로그 동시 기록."""
    print(f"[xlsx] {msg}", flush=True)
    if job_id is None:
        return
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job:
            job["logs"].append({"t": _time.time() - job["created"], "msg": msg})


def _cleanup_old_jobs():
    now = _time.time()
    with _jobs_lock:
        stale = [jid for jid, j in _jobs.items() if now - j["created"] > JOB_TTL_SECONDS]
        for jid in stale:
            j = _jobs.pop(jid, None)
            if j and j.get("file_path") and os.path.exists(j["file_path"]):
                try:
                    os.unlink(j["file_path"])
                except OSError:
                    pass


def _resolve_date_range(since, until, date_preset):
    if since and until:
        return since, until
    preset = date_preset or "last_7d"
    today = date.today()
    if preset == "today":
        s, u = today, today
    elif preset == "yesterday":
        s = u = today - timedelta(days=1)
    elif preset == "last_7d":
        s, u = today - timedelta(days=7), today - timedelta(days=1)
    elif preset == "last_14d":
        s, u = today - timedelta(days=14), today - timedelta(days=1)
    elif preset == "last_30d":
        s, u = today - timedelta(days=30), today - timedelta(days=1)
    elif preset == "this_month":
        s, u = today.replace(day=1), today
    elif preset == "last_month":
        first_this = today.replace(day=1)
        u = first_this - timedelta(days=1)
        s = u.replace(day=1)
    else:
        s, u = today - timedelta(days=7), today - timedelta(days=1)
    return s.isoformat(), u.isoformat()


def _run_xlsx_job(job_id, resolved, tpl, since, until):
    """별도 스레드에서 실행되는 워커. 진행 상황을 _emit으로 기록."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = "openpyxl이 설치되지 않았습니다. pip install -r requirements.txt 후 재시작하세요."
        return

    t0 = _time.time()
    try:
        _emit(job_id, f"Meta API 호출 시작: {since} ~ {until}")
        result = _fetch_template_insights_with_progress(resolved, since, until, job_id)
        if "error" in result:
            with _jobs_lock:
                _jobs[job_id]["status"] = "error"
                _jobs[job_id]["error"] = result["error"]
            return

        rows = result["data"]
        _emit(job_id, f"API 응답 완료: 총 {len(rows)}행 (+{_time.time()-t0:.1f}s)")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy(tpl["template_path"], tmp.name)
        _emit(job_id, f"템플릿 복사 완료 (+{_time.time()-t0:.1f}s)")

        _emit(job_id, "워크북 로드 중... (25MB 파일이라 20~40초 소요)")
        wb = load_workbook(tmp.name, keep_links=False)
        _emit(job_id, f"워크북 로드 완료 (+{_time.time()-t0:.1f}s)")

        if tpl["raw_sheet"] not in wb.sheetnames:
            raise RuntimeError(f"템플릿에 '{tpl['raw_sheet']}' 시트가 없습니다.")
        ws = wb[tpl["raw_sheet"]]

        cols = tpl["columns"]
        header_row = tpl["header_row"]
        start_row = header_row + 1
        K_COL = 11

        existing_max = ws.max_row
        new_count = len(rows)
        new_max = start_row + new_count - 1
        _emit(job_id, f"기존 데이터: {max(0, existing_max - header_row)}행 / 신규: {new_count}행")

        for i, row in enumerate(rows):
            target_row = start_row + i
            actions = row.get("actions", [])
            result_type, result_count = _summarize_result(actions)
            video_3sec = _summarize_video_3sec(actions)

            ws.cell(row=target_row, column=cols["date"], value=row.get("date_start", ""))
            ws.cell(row=target_row, column=cols["campaign_name"], value=row.get("campaign_name", ""))
            ws.cell(row=target_row, column=cols["adset_name"], value=row.get("adset_name", ""))
            ws.cell(row=target_row, column=cols["ad_name"], value=row.get("ad_name", ""))
            ws.cell(row=target_row, column=cols["impressions"], value=_to_float(row.get("impressions")))
            ws.cell(row=target_row, column=cols["link_clicks"], value=_to_float(row.get("inline_link_clicks")))
            ws.cell(row=target_row, column=cols["video_3sec"], value=video_3sec)
            ws.cell(row=target_row, column=cols["result_type"], value=result_type)
            ws.cell(row=target_row, column=cols["result_count"], value=result_count)
            ws.cell(row=target_row, column=cols["spend"], value=_to_float(row.get("spend")))

        if existing_max > new_max:
            cols_to_clear = list(cols.values()) + [K_COL]
            for r in range(new_max + 1, existing_max + 1):
                for c in cols_to_clear:
                    ws.cell(row=r, column=c).value = None
            _emit(job_id, f"잉여 {existing_max - new_max}행 클리어")

        if new_max > existing_max:
            for r in range(existing_max + 1, new_max + 1):
                ws.cell(row=r, column=K_COL,
                        value=f"=VLOOKUP(D{r},raw_table!$E$2:$F$1000,2,FALSE)")

        _emit(job_id, f"셀 쓰기 완료 (+{_time.time()-t0:.1f}s) — 저장 시작 (10~30초 소요)")
        wb.save(tmp.name)
        _emit(job_id, f"✓ 저장 완료, 총 {_time.time()-t0:.1f}s")

        label = tpl["label"].replace(" ", "_")
        filename = f"{label}_report_{since}_to_{until}.xlsx"

        with _jobs_lock:
            _jobs[job_id]["file_path"] = tmp.name
            _jobs[job_id]["filename"] = filename
            _jobs[job_id]["status"] = "done"
    except Exception as e:
        _emit(job_id, f"✗ 오류: {e}")
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = str(e)


def _fetch_template_insights_with_progress(account_id, since, until, job_id):
    """페이지네이션 진행 상황을 잡 로그로 흘려보내는 버전."""
    template_fields = [
        "date_start", "campaign_name", "adset_name", "ad_name",
        "impressions", "inline_link_clicks", "spend", "actions",
    ]
    if not META_ACCESS_TOKEN:
        return {"error": "META_ACCESS_TOKEN이 비어 있습니다."}

    url = f"{GRAPH_URL}/{account_id}/insights"
    params = {
        "fields": ",".join(template_fields),
        "level": "ad",
        "time_increment": 1,
        "time_range": json.dumps({"since": since, "until": until}),
        "access_token": META_ACCESS_TOKEN,
        "limit": 500,
    }

    all_rows = []
    next_url, next_params = url, params
    pages = 0
    while next_url and pages < 50:
        pages += 1
        try:
            r = requests.get(next_url, params=next_params, timeout=60)
            data = r.json()
        except requests.RequestException as e:
            return {"error": f"네트워크 오류: {e}"}
        if "error" in data:
            return {"error": data["error"].get("message", "Meta API 오류")}
        batch = data.get("data", [])
        all_rows.extend(batch)
        _emit(job_id, f"  페이지 {pages} 수신: +{len(batch)}행 (누적 {len(all_rows)}행)")
        paging = data.get("paging", {})
        next_url = paging.get("next")
        next_params = None
    return {"data": all_rows}


@app.route("/api/report.xlsx/start", methods=["POST"])
def api_report_start():
    if not META_ACCESS_TOKEN:
        return jsonify({"error": "META_ACCESS_TOKEN이 비어 있습니다."}), 400

    account_id_req = request.args.get("account_id")
    resolved = _resolve_account_id(account_id_req)
    if not resolved:
        return jsonify({"error": "허용되지 않은 광고 계정 ID입니다."}), 400

    tpl = ACCOUNT_TEMPLATES.get(resolved)
    if not tpl:
        return jsonify({"error": f"이 계정({resolved})은 템플릿이 등록되지 않았습니다. 현재는 BC 계정만 지원합니다."}), 400
    if not os.path.exists(tpl["template_path"]):
        return jsonify({"error": f"템플릿 파일을 찾을 수 없습니다: {tpl['template_path']}"}), 500

    since, until = _resolve_date_range(
        request.args.get("since"),
        request.args.get("until"),
        request.args.get("date_preset"),
    )

    _cleanup_old_jobs()
    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "running",
            "logs": [],
            "file_path": None,
            "filename": None,
            "error": None,
            "created": _time.time(),
        }

    threading.Thread(
        target=_run_xlsx_job,
        args=(job_id, resolved, tpl, since, until),
        daemon=True,
    ).start()

    return jsonify({"job_id": job_id, "since": since, "until": until})


@app.route("/api/report.xlsx/stream/<job_id>")
def api_report_stream(job_id):
    """Server-Sent Events: 잡 진행 로그를 실시간 푸시."""
    def event_stream():
        sent_logs = 0
        # Some browsers buffer until first chunk — send initial comment
        yield ": stream-open\n\n"
        while True:
            with _jobs_lock:
                job = _jobs.get(job_id)
                if not job:
                    yield f"event: error\ndata: {json.dumps({'msg': 'job not found'})}\n\n"
                    return
                logs = list(job["logs"])
                status = job["status"]
                error = job.get("error")

            new_logs = logs[sent_logs:]
            sent_logs = len(logs)
            for log in new_logs:
                yield f"data: {json.dumps(log, ensure_ascii=False)}\n\n"

            if status == "done":
                yield f"event: done\ndata: {json.dumps({'job_id': job_id})}\n\n"
                return
            if status == "error":
                yield f"event: failed\ndata: {json.dumps({'msg': error or '알 수 없는 오류'}, ensure_ascii=False)}\n\n"
                return
            _time.sleep(0.4)

    return Response(
        event_stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.route("/api/report.xlsx/download/<job_id>")
def api_report_download(job_id):
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return "Job not found", 404
        if job["status"] != "done":
            return f"Job not ready (status={job['status']})", 400
        file_path = job["file_path"]
        filename = job["filename"]
    if not file_path or not os.path.exists(file_path):
        return "File missing", 500
    return send_file(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _agg_metrics_full(rows):
    """프론트 aggregateMetrics와 동일. video_views는 모두 null이면 None."""
    imps = sum(_to_float(r.get("impressions")) for r in rows)
    clicks = sum(_to_float(r.get("clicks")) for r in rows)
    spend = sum(_to_float(r.get("spend")) for r in rows)
    conv = sum(_to_float(r.get("conversions") or 0) for r in rows)
    has_video = any(r.get("video_views") is not None for r in rows)
    if has_video:
        views = sum(_to_float(r.get("video_views") or 0) for r in rows if r.get("video_views") is not None)
    else:
        views = None
    return {
        "impressions": imps,
        "clicks": clicks,
        "ctr": (clicks / imps * 100) if imps else 0.0,
        "video_views": int(views) if (views is not None) else None,
        "vtr": (views / imps * 100) if (views is not None and imps) else None,
        "conversions": int(conv),
        "cvr": (conv / imps * 100) if imps else 0.0,
        "spend": spend,
    }


def _dow_kr(date_str):
    """YYYY-MM-DD → 한글 요일(월/화/.../일)."""
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return ""
    return ["월", "화", "수", "목", "금", "토", "일"][d.weekday()]


def _safe_sheet_name(name):
    """Excel 시트 이름 sanitize: 최대 31자, 금지문자 제거."""
    if not name:
        return "Sheet"
    bad = set('/\\?*:[]')
    out = "".join("_" if c in bad else c for c in str(name))
    return out[:31] or "Sheet"


@app.route("/api/view.xlsx")
def api_view_xlsx():
    """현재 뷰 그대로 단순 xlsx 출력. 전체 본문 try/except로 감싸 에러 추적 강화."""
    try:
        return _build_view_xlsx_response()
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[view.xlsx] FATAL: {e}\n{tb}", flush=True)
        return jsonify({
            "error": f"Excel 생성 실패: {e}",
            "trace_tail": tb.splitlines()[-10:],
        }), 500


def _build_view_xlsx_response():
    """원래 api_view_xlsx 본문. 새 시트 구조: 주간 인사이트 + SUMMARY + 일자별_[Target] per Target."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl이 설치되지 않았습니다. pip install -r requirements.txt 후 재시작하세요."}), 500

    resolved = _resolve_account_id(request.args.get("account_id"))
    if not resolved:
        return jsonify({"error": "허용되지 않은 광고 계정 ID입니다."}), 400

    since, until = _resolve_date_range(
        request.args.get("since"),
        request.args.get("until"),
        request.args.get("date_preset"),
    )
    level = request.args.get("level", "campaign")
    overrides = _parse_overrides(request.args.get("override", ""))

    wb = Workbook()
    base_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1877F2")
    hatched_fill = PatternFill(patternType="lightUp", fgColor="C0C0C0", bgColor="F0F0F0")
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    metric_keys = ["impressions", "clicks", "ctr", "video_views", "vtr", "conversions", "cvr", "spend"]
    metric_labels = ["노출", "클릭", "CTR", "VIEW", "VTR", "전환", "CVR", "지출(원)"]
    metric_fmts = {
        "impressions": "#,##0", "clicks": "#,##0", "ctr": "0.00%",
        "video_views": "#,##0", "vtr": "0.00%",
        "conversions": "#,##0", "cvr": "0.00%", "spend": '#,##0"원"',
    }

    def write_metric(cell, val, key):
        cell.font = base_font
        cell.border = border
        if val is None and key in ("video_views", "vtr"):
            cell.fill = hatched_fill
            return
        if val is None:
            return
        if key in ("ctr", "vtr", "cvr"):
            cell.value = (val or 0) / 100.0
        else:
            cell.value = val
        if key in metric_fmts:
            cell.number_format = metric_fmts[key]
        cell.alignment = Alignment(horizontal="right")

    def write_header_row(ws, headers, row=1):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")

    # === Sheet 1: 주간 인사이트 ===
    ws1 = wb.active
    ws1.title = "주간 인사이트"
    this_start, this_end, last_start, last_end = _get_week_ranges()
    this_res = fetch_insights(since=this_start.isoformat(), until=this_end.isoformat(), level="campaign", account_id=resolved)
    last_res = fetch_insights(since=last_start.isoformat(), until=last_end.isoformat(), level="campaign", account_id=resolved)

    if "error" not in this_res and "error" not in last_res:
        this_total = _aggregate(this_res["data"])
        last_total = _aggregate(last_res["data"])
        deltas = _build_deltas(this_total, last_total)
        up, down = _build_movers(this_res["data"], last_res["data"])
        period = {
            "this": {"start": this_start.isoformat(), "end": this_end.isoformat()},
            "last": {"start": last_start.isoformat(), "end": last_end.isoformat()},
            "days": (this_end - this_start).days + 1,
        }
        bullets = _build_bullets(this_total, last_total, deltas, up, down, period)

        ws1.cell(row=1, column=1, value="📊 주간 인사이트").font = Font(name="Arial", size=14, bold=True)
        ws1.cell(row=2, column=1, value=f"이번 주 {period['this']['start']}~{period['this']['end']} ({period['days']}일) vs 지난 주 {period['last']['start']}~{period['last']['end']}").font = Font(name="Arial", size=10, italic=True, color="666666")

        write_header_row(ws1, ["항목", "이번 주", "지난 주", "변동", "변동률"], row=4)
        rows_data = [
            ("총 지출", deltas["spend"]["this"], deltas["spend"]["last"], deltas["spend"]["abs"], deltas["spend"]["pct"]),
            ("노출", deltas["impressions"]["this"], deltas["impressions"]["last"], deltas["impressions"]["abs"], deltas["impressions"]["pct"]),
            ("클릭", deltas["clicks"]["this"], deltas["clicks"]["last"], deltas["clicks"]["abs"], deltas["clicks"]["pct"]),
            ("CTR (%)", deltas["ctr"]["this"], deltas["ctr"]["last"], deltas["ctr"]["abs"], deltas["ctr"]["pct"]),
            ("CPC (원)", deltas["cpc"]["this"], deltas["cpc"]["last"], deltas["cpc"]["abs"], deltas["cpc"]["pct"]),
            ("CPM (원)", deltas["cpm"]["this"], deltas["cpm"]["last"], deltas["cpm"]["abs"], deltas["cpm"]["pct"]),
        ]
        for i, (label, this_v, last_v, abs_v, pct_v) in enumerate(rows_data, start=5):
            ws1.cell(row=i, column=1, value=label).font = base_font
            for col_idx, val in enumerate([this_v, last_v, abs_v], start=2):
                c = ws1.cell(row=i, column=col_idx, value=val)
                c.font = base_font
                c.number_format = "#,##0.00" if "%" in label or "원" in label else "#,##0"
                c.alignment = Alignment(horizontal="right")
            pct_cell = ws1.cell(row=i, column=5, value=(pct_v / 100.0) if pct_v is not None else None)
            pct_cell.font = base_font
            pct_cell.number_format = "0.0%;-0.0%;-"
            pct_cell.alignment = Alignment(horizontal="right")

        # bullets
        bullets_start_row = 5 + len(rows_data) + 2
        ws1.cell(row=bullets_start_row, column=1, value="💡 자동 인사이트").font = Font(name="Arial", size=11, bold=True)
        for i, b in enumerate(bullets, start=bullets_start_row + 1):
            ws1.cell(row=i, column=1, value="• " + b).font = base_font
            ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

        for col, w in [("A", 18), ("B", 16), ("C", 16), ("D", 16), ("E", 12)]:
            ws1.column_dimensions[col].width = w

    # === 광고 단위 일별 데이터 fetch (모든 후속 시트 공용) ===
    period_data = fetch_insights(since=since, until=until, level="ad", account_id=resolved)
    if "error" in period_data:
        # 에러 시 SUMMARY 시트 하나만 만들고 에러 메시지 기록
        ws_err = wb.create_sheet("ERROR")
        ws_err.cell(row=1, column=1, value=f"데이터 로드 실패: {period_data['error']}").font = Font(name="Arial", color="C0392B", bold=True)
    else:
        all_rows = period_data["data"]
        # 모든 행에 _target/_targeting/_targeting_main/_creative 부여
        for r in all_rows:
            _annotate_row_for_export(r)

        # ===== Sheet 2: SUMMARY (Target × Targeting × Creative) =====
        ws2 = wb.create_sheet("SUMMARY")
        bold_font = Font(name="Arial", size=10, bold=True)
        white_bold = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        targeting_fill = PatternFill("solid", fgColor="EEF3F8")
        targeting_subtotal_fill = PatternFill("solid", fgColor="E8F1FD")
        target_subtotal_fill = PatternFill("solid", fgColor="D6E5FA")
        total_fill = PatternFill("solid", fgColor="5B6C7E")

        write_header_row(ws2, ["Target", "Targeting", "Creative"] + metric_labels, row=1)

        # 그룹: Target → Targeting → {rows, creatives: {short_label: {rows, full}}}
        grouped = {}
        for r in all_rows:
            t = r.get("_target") or "기타"
            tg = r.get("_targeting") or "기타"
            cr = r.get("_creative") or (r.get("ad_name") or "(이름 없음)")
            full = r.get("ad_name") or cr
            grouped.setdefault(t, {})
            if tg not in grouped[t]:
                grouped[t][tg] = {"rows": [], "creatives": {}}
            grouped[t][tg]["rows"].append(r)
            if cr not in grouped[t][tg]["creatives"]:
                grouped[t][tg]["creatives"][cr] = {"rows": [], "full": full}
            grouped[t][tg]["creatives"][cr]["rows"].append(r)

        target_order = ["Adults", "Young Learners", "Kinder"]
        def _t_sort_key(t):
            try:
                return (target_order.index(t), t)
            except ValueError:
                return (999, t)
        sorted_targets = sorted(grouped.keys(), key=_t_sort_key)

        try:
            from openpyxl.comments import Comment
        except ImportError:
            Comment = None

        row_idx = 2
        for target in sorted_targets:
            targeting_map = grouped[target]
            tg_keys = sorted(targeting_map.keys())
            target_first_row = row_idx

            for tg in tg_keys:
                entry = targeting_map[tg]
                tg_metrics = _agg_metrics_full(entry["rows"])

                # Targeting 헤더 행 — 메트릭은 Targeting 합계
                ws2.cell(row=row_idx, column=1, value=(target if row_idx == target_first_row else "")).font = base_font
                ws2.cell(row=row_idx, column=2, value=tg).font = bold_font
                ws2.cell(row=row_idx, column=3, value="")
                for i, key in enumerate(metric_keys):
                    write_metric(ws2.cell(row=row_idx, column=4 + i), tg_metrics.get(key), key)
                for col in range(1, 4 + len(metric_keys)):
                    ws2.cell(row=row_idx, column=col).fill = targeting_fill
                row_idx += 1

                # Creative 행들 (지출 큰 순)
                cr_entries = []
                for cr_name, cr_data in entry["creatives"].items():
                    cm = _agg_metrics_full(cr_data["rows"])
                    cr_entries.append((cr_name, cr_data["full"], cm))
                cr_entries.sort(key=lambda x: -(x[2].get("spend") or 0))

                for cr_name, cr_full, cm in cr_entries:
                    name_cell = ws2.cell(row=row_idx, column=3, value=cr_name)
                    name_cell.font = base_font
                    if Comment and cr_full and cr_full != cr_name:
                        name_cell.comment = Comment(cr_full, "Auto")
                    for i, key in enumerate(metric_keys):
                        write_metric(ws2.cell(row=row_idx, column=4 + i), cm.get(key), key)
                    row_idx += 1

                # Sub Total per Targeting
                ws2.cell(row=row_idx, column=3, value=f"Sub Total ({tg})").font = bold_font
                for i, key in enumerate(metric_keys):
                    write_metric(ws2.cell(row=row_idx, column=4 + i), tg_metrics.get(key), key)
                for col in range(1, 4 + len(metric_keys)):
                    ws2.cell(row=row_idx, column=col).fill = targeting_subtotal_fill
                    ws2.cell(row=row_idx, column=col).font = bold_font
                row_idx += 1

            # Sub Total per Target
            target_all_rows = [r for tg_data in targeting_map.values() for r in tg_data["rows"]]
            target_metrics = _agg_metrics_full(target_all_rows)
            ws2.cell(row=row_idx, column=2, value=f"Sub Total ({target})").font = bold_font
            for i, key in enumerate(metric_keys):
                write_metric(ws2.cell(row=row_idx, column=4 + i), target_metrics.get(key), key)
            for col in range(1, 4 + len(metric_keys)):
                ws2.cell(row=row_idx, column=col).fill = target_subtotal_fill
                ws2.cell(row=row_idx, column=col).font = bold_font
            row_idx += 1

        # 전체 Total
        total_metrics = _agg_metrics_full(all_rows)
        ws2.cell(row=row_idx, column=1, value="Total").font = white_bold
        for i, key in enumerate(metric_keys):
            write_metric(ws2.cell(row=row_idx, column=4 + i), total_metrics.get(key), key)
        for col in range(1, 4 + len(metric_keys)):
            ws2.cell(row=row_idx, column=col).fill = total_fill
            ws2.cell(row=row_idx, column=col).font = white_bold

        # 컬럼 너비
        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 36
        for i in range(4, 4 + len(metric_keys)):
            ws2.column_dimensions[get_column_letter(i)].width = 12

        # ===== Sheet 3+: 일자별_[Target] per Target (stacked panels) =====
        # 공통 fills
        group_header_fill = PatternFill("solid", fgColor="5B6C7E")
        metric_header_fill = PatternFill("solid", fgColor="7A8A9C")
        date_total_fill = PatternFill("solid", fgColor="FFF3A8")
        weekend_red = Font(name="Arial", size=10, color="C0392B")

        all_dates = sorted({r.get("date_start") for r in all_rows if r.get("date_start")})

        for target in sorted_targets:
            target_rows = [r for r in all_rows if r.get("_target") == target]
            if not target_rows:
                continue
            sheet_name = _safe_sheet_name(f"일자별_{target}")
            wsd = wb.create_sheet(sheet_name)

            r_idx = 1

            # Panel A: Date × [Target Total + 각 Targeting]
            targetings_in = sorted({r.get("_targeting") for r in target_rows if r.get("_targeting")})
            groups_a = [{"label": f"{target} Total", "full": None, "rows": target_rows}]
            for tg in targetings_in:
                groups_a.append({
                    "label": tg,
                    "full": None,
                    "rows": [r for r in target_rows if r.get("_targeting") == tg],
                })
            r_idx = _write_daily_panel(
                wsd, r_idx,
                f"📅 Daily by Targeting — {target}",
                groups_a, all_dates,
                base_font, bold_font, write_metric,
                group_header_fill, metric_header_fill, date_total_fill, weekend_red,
                Comment,
            )
            r_idx += 2  # 빈 행 2개

            # Panel B per Targeting: Date × [Targeting Total + 각 Creative]
            for tg in targetings_in:
                tg_rows = [r for r in target_rows if r.get("_targeting") == tg]
                # Creative 추출
                creative_map = {}  # short → {rows, full}
                for r in tg_rows:
                    cr = r.get("_creative") or (r.get("ad_name") or "(이름 없음)")
                    full = r.get("ad_name") or cr
                    if cr not in creative_map:
                        creative_map[cr] = {"rows": [], "full": full}
                    creative_map[cr]["rows"].append(r)
                creative_keys = sorted(creative_map.keys(),
                                       key=lambda c: -(_agg_metrics_full(creative_map[c]["rows"]).get("spend") or 0))

                groups_b = [{"label": f"{tg} Total", "full": None, "rows": tg_rows}]
                for cr in creative_keys:
                    groups_b.append({
                        "label": cr,
                        "full": creative_map[cr]["full"],
                        "rows": creative_map[cr]["rows"],
                    })
                r_idx = _write_daily_panel(
                    wsd, r_idx,
                    f"📅 Daily by Creative — {target} > {tg}",
                    groups_b, all_dates,
                    base_font, bold_font, write_metric,
                    group_header_fill, metric_header_fill, date_total_fill, weekend_red,
                    Comment,
                )
                r_idx += 2

            # 컬럼 너비
            wsd.column_dimensions['A'].width = 12
            wsd.column_dimensions['B'].width = 6

    # 저장 & 응답
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)

    tpl = ACCOUNT_TEMPLATES.get(resolved)
    label = (tpl["label"] if tpl else resolved).replace(" ", "_")
    filename = f"{label}_view_{since}_to_{until}.xlsx"

    return send_file(
        tmp.name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _write_daily_panel(ws, start_row, title, groups, dates,
                       base_font, bold_font, write_metric,
                       group_header_fill, metric_header_fill, date_total_fill, weekend_red,
                       Comment):
    """일자별 데이터 패널 1개 작성. Date × [Group Total + 각 sub-group] 형태. 반환: 다음 사용 가능 row."""
    from openpyxl.styles import Alignment, Font, PatternFill
    metric_keys = ["impressions", "clicks", "ctr", "video_views", "vtr", "conversions", "cvr", "spend"]
    metric_labels = ["IMPs", "Click", "CTR", "View", "VTR", "전환", "CVR", "Spend"]
    M = len(metric_keys)
    white_bold = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    r = start_row

    # 패널 타이틀
    title_cell = ws.cell(row=r, column=1, value=title)
    title_cell.font = Font(name="Arial", size=11, bold=True)
    r += 1

    # 그룹 헤더 행 (Date, 요일은 rowspan 2, 그룹 라벨은 colspan M)
    ws.cell(row=r, column=1, value="Date").font = white_bold
    ws.cell(row=r, column=1).fill = group_header_fill
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)

    ws.cell(row=r, column=2, value="요일").font = white_bold
    ws.cell(row=r, column=2).fill = group_header_fill
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=2)

    col = 3
    for g in groups:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + M - 1)
        cell = ws.cell(row=r, column=col, value=g["label"])
        cell.font = white_bold
        cell.fill = group_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if Comment and g.get("full") and g["full"] != g["label"]:
            cell.comment = Comment(g["full"], "Auto")
        col += M
    r += 1

    # 메트릭 라벨 행
    col = 3
    for g in groups:
        for lbl in metric_labels:
            cell = ws.cell(row=r, column=col, value=lbl)
            cell.font = white_bold
            cell.fill = metric_header_fill
            cell.alignment = Alignment(horizontal="center")
            col += 1
    r += 1

    # Total 행 (전체 기간 합계)
    ws.cell(row=r, column=1, value="Total").font = bold_font
    ws.cell(row=r, column=1).fill = date_total_fill
    ws.cell(row=r, column=2, value="")
    ws.cell(row=r, column=2).fill = date_total_fill
    col = 3
    for g in groups:
        m = _agg_metrics_full(g["rows"])
        for key in metric_keys:
            write_metric(ws.cell(row=r, column=col), m.get(key), key)
            ws.cell(row=r, column=col).fill = date_total_fill
            ws.cell(row=r, column=col).font = bold_font
            col += 1
    r += 1

    # 날짜 행들
    for d in dates:
        dow = _dow_kr(d)
        ws.cell(row=r, column=1, value=d)
        ws.cell(row=r, column=2, value=dow)
        if dow in ("토", "일"):
            ws.cell(row=r, column=1).font = weekend_red
            ws.cell(row=r, column=2).font = weekend_red
        col = 3
        for g in groups:
            date_rows = [row for row in g["rows"] if row.get("date_start") == d]
            if date_rows:
                m = _agg_metrics_full(date_rows)
                for key in metric_keys:
                    write_metric(ws.cell(row=r, column=col), m.get(key), key)
                    col += 1
            else:
                col += M
        r += 1

    return r


@app.route("/api/insights.csv")
def api_insights_csv():
    result = fetch_insights(
        date_preset=request.args.get("date_preset"),
        since=request.args.get("since"),
        until=request.args.get("until"),
        level=request.args.get("level", "campaign"),
        account_id=request.args.get("account_id"),
    )
    if "error" in result:
        return result["error"], 400

    rows = result["data"]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in CSV_COLUMNS})

    csv_data = "﻿" + buf.getvalue()
    filename = f"meta_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
